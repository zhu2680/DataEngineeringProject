import logging

import azure.functions as func
import azure.storage.blob
import io
import pdfplumber
from openpyxl import Workbook, load_workbook
from string import ascii_uppercase as alp 

def main(myblob: func.InputStream):
    logging.info(f"Python blob trigger function processed blob \n"
                 f"Name: {myblob.name}\n"
                 f"Blob Size: {myblob.length} bytes")

    #formatting of this allows you too call it from powershell 
   
    #this will be our row counter 
    count = 1
    #build array to call columns from
    columns = []
    for i in alp: 
        columns.append(i)
    for i in alp: 
        for j in alp: 
            columns.append(i+j)
    logging.info('Loading pdf...')
    #for filename in os.listdir(pdf_folder):
    if myblob.name.endswith('.pdf'): 
        #this is error mitigation so even if it fails to read a pdf it will let you know if it couldnt read it sometimes this is due to corupt files
        # if will try this section of code and if it fails it will go to the EXCEPTION on line 330
        try:
            #pdf_path= os.path.join(pdf_folder, filename)
            pdf_path = io.BytesIO(myblob.read())
            text = scrape_pdf_table(pdf_path)
            logging.info("finish scraping the pdf table")
            # write to a excel file
            test = Workbook()
            current0 = test.active
            lines = text.split("\n")
            for row, line in enumerate(lines, start=1):
                columns = line.split(" ")
                for col, column in enumerate(columns, start=1):
                    current0.cell(row=row, column=col, value=column)
            logging.info(f"here is the value in A1: {current0['A1'].value}")
            logging.info("finish converting the pdf to excel")
            
            conString = "DefaultEndpointsProtocol=https;AccountName=mrtbartonabb3a1607b;AccountKey=awSIfw1YfTIclurJu6hrYrjjbgNMK1cTliTnEUv6VfFVKCfrmFAvzYx/m6ccV2ZnfBw2SaKORipp+AStQh5dwQ==;EndpointSuffix=core.windows.net"
            ##"DefaultEndpointsProtocol=https;AccountName=mrtbartonabb3a1607b;AccountKey=eoJHu+6vJYYXEiv59LuHSouVaRxVL/XRcXkpVBQu4oVnKUY25Q1ZKgPb7LgbC/ax5jYTY5/r0DEY+AStkH/00w==;EndpointSuffix=core.windows.net"
            blob_service_client = azure.storage.blob.BlobServiceClient.from_connection_string(conString)
            excelname = myblob.name[15:]
            
            blob_client = blob_service_client.get_blob_client(container='inspectionformexcel', blob = excelname[:-4]+".xlsx")
            
            output_file = io.BytesIO()
            test.save(output_file)
            blob_client.upload_blob(output_file.getvalue())    
            logging.info("excel version is saved to container")
            #excel_path = os.path.join(pdf_folder, f"{filename[:-4]}.xlsx")
            #write_to_excel(text, excel_path)
            #wb = load_workbook(excel_path)
            #this is our current pdf we have scraped
            #current = wb.active
            
            #pass     # if the above code passes it will call this code 
            
            
            

        except Exception as e: 
            logging.error(f"Unable to read file {myblob.name} an error occured: {e}")
            
        


def scrape_pdf_table(file_path):
    with pdfplumber.open(file_path) as pdf:
        text=""
        for page in pdf.pages: 
            text += page.extract_text()
    return text





