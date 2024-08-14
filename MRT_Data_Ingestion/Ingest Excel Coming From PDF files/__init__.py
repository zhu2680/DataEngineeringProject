import logging
import azure.functions as func
import azure.storage.blob
import io

from openpyxl import load_workbook
from string import ascii_uppercase as alp 
import azure.functions as func


def main(myblob: func.InputStream):
    logging.info(f"Python blob trigger function processed blob \n"
                 f"Name: {myblob.name}\n"
                 f"Blob Size: {myblob.length} bytes")

    #count = 1
    #build array to call columns from
    columns = []
    for i in alp: 
        columns.append(i)
    for i in alp: 
        for j in alp: 
            columns.append(i+j)
    logging.info('Loading pdf...')
    wb = load_workbook(filename = io.BytesIO(myblob.read()), data_only= True)
    current = wb.worksheets[0]

    logging.info("the scraped excel is loaded")
    title = current['A1'].value
    if title == "Inspection":
        try:
            
            maxrows = count_max(current)

            # this is our master file where we will write all info too
            #test = load_workbook(r"C:\dev\mrt\PDF_MRT.xlsx", read_only = False)
            #full_table = test.active
            
            
            
            
            conString = "DefaultEndpointsProtocol=https;AccountName=mrtbartonabb3a1607b;AccountKey=awSIfw1YfTIclurJu6hrYrjjbgNMK1cTliTnEUv6VfFVKCfrmFAvzYx/m6ccV2ZnfBw2SaKORipp+AStQh5dwQ==;EndpointSuffix=core.windows.net"
            ##"DefaultEndpointsProtocol=https;AccountName=mrtbartonabb3a1607b;AccountKey=eoJHu+6vJYYXEiv59LuHSouVaRxVL/XRcXkpVBQu4oVnKUY25Q1ZKgPb7LgbC/ax5jYTY5/r0DEY+AStkH/00w==;EndpointSuffix=core.windows.net"
            blob_service_client = azure.storage.blob.BlobServiceClient.from_connection_string(conString)
            
            
            #full_table = result.worksheets[0]
            blob_client = blob_service_client.get_blob_client(container='inspectionformscraped', blob = "MasterInspection.xlsx")
            with io.BytesIO() as input_blob:
                blob_client.download_blob().readinto(input_blob)
                input_blob.seek(0)
                result = load_workbook(filename = input_blob, data_only= True)
            full_table= result.worksheets[0]

            logging.info("finish loading the master excel")
            count = count_max(full_table)
            logging.info(f"The max number of non-empty row is: {count}")


            count += 1 
            dict = build_dictionary(current,columns)
            logging.info("finish creating the title dic")
            dict2 = build_dictionary_as(current,maxrows,columns)

            logging.info("finished creating the dic for tables, start writing to the result excel...")
            full_table['A'+ str(count)].value= myblob.name[:-4]
            for num, item in enumerate(dict, start=1):
                full_table[columns[num]+ str(count)].value = dict[item]
            if "afdvolume" in dict2:
                full_table[columns[num+1]+str(count)].value = dict2["afdvolume"]
            if "afdenergy" in dict2:
                full_table[columns[num+2]+str(count)].value = dict2["afdenergy"]
            if "afdvolume" in dict2:
                full_table[columns[num+3]+str(count)].value = dict2["aftolerance"]
            if "afdenergy" in dict2:
                full_table[columns[num+4]+str(count)].value = dict2["afpass/fail"]
            
            if "aldvolume" in dict2:
                full_table[columns[num+5]+str(count)].value = dict2["aldvolume"]
            if "aldenergy" in dict2:
                full_table[columns[num+6]+str(count)].value = dict2["aldenergy"]
            if "aldvolume" in dict2:
                full_table[columns[num+7]+str(count)].value = dict2["altolerance"]
            if "aldenergy" in dict2:
                full_table[columns[num+8]+str(count)].value = dict2["alpass/fail"]
            
            #test.save(r"C:\dev\mrt\PDF_MRT.xlsx")
            logging.info("finshed scraping the info needed from excel")
            
            
            # test.save(tmp.name)
            # output = io.BytesIO(tmp.read())
            # blob_client.upload_blob(output)
            output = io.BytesIO()
            result.save(output)
            blob_client.upload_blob(output.getvalue(), overwrite=True)
            logging.info(f"{myblob.name} has been read :)")
        except Exception as e: 
            logging.error(f"Unable to read file {myblob.name} an error occured: {e}")

#checks inconsistancys 
def build_dictionary(current,columns):
    #check if location name is two names
    dict_table = {}
    #SITE INFORMATION
    dict_table["date"] = current['E1'].value
    #find location and run number
    b = 1
    a = columns[b]
    loc = current[a+'3'].value
    while current[columns[b+1]+'3'].value != "Run":
        b += 1
        a = columns[b]
        loc += current[a+'3'].value
    
    dict_table['loc']= loc
    
    b3 = b+3
    a3 = columns[b3]
    run_num = current[a3+'3'].value
    while current[columns[b3+1]+'3'].value != None:
        b3 += 1
        a3 = columns[b3]
        #print(current[a3+'3'].value)
        #print("hello")
        run_num += current[a3+'3'].value
        #print(current[a3+'3'].value)
    dict_table['run_num']= run_num
    
    #find station name and station number
    b1 = 2
    a1 = columns[b1]
    stn = current[a1+'4'].value
    while current[columns[b1+1]+'4'].value != "Stn.":
        b1 += 1
        a1 = columns[b1]
        stn += current[a1+'4'].value
    
    dict_table['stn_name']= stn
    dict_table['stn_num']= current[columns[b1+3]+'4'].value
    dict_table['tec_name']= current['B5'].value + " " + current['C5'].value 
    return dict_table


def build_dictionary_as(current,maxrows,columns):    
    dict_table = {}
    #test result as found
    af_col, af_row = searchstring(65, "Test", "Results", current, maxrows, columns)
    
    while current[columns[af_col+1]+str(af_row)].value != None:
        af_col +=1
    result = current[columns[af_col]+str(af_row)].value
    
    
    lastd = check_as(result)
    print(lastd)

     #as found
    diffV_col,diffV_row = searchstring(af_row, "Difference,", "Volume:", current, maxrows, columns)
    
    
    #if lastd == "al":
    dict_table[lastd+'dvolume'] = current[columns[diffV_col]+str(diffV_row)].value
    dict_table[lastd+'denergy'] = current[columns[diffV_col]+str(diffV_row+1)].value
    dict_table[lastd+'tolerance'] = current[columns[diffV_col-1]+str(diffV_row+2)].value
    dict_table[lastd+'pass/fail'] = current[columns[diffV_col-2]+str(diffV_row+3)].value
        
    #test result as left if there exists a third page
    if maxrows >= diffV_row + 20:
        af_col1, af_row1 = searchstring(af_row+35, "Test", "Results", current, maxrows, columns)
        
        while current[columns[af_col1+1]+str(af_row1)].value != None:
            af_col1 +=1
        result2 = current[columns[af_col1]+str(af_row1)].value
        lastd2 = check_as(result2)
        diffV_col1,diffV_row1 = searchstring(af_row1, "Difference,", "Volume:", current, maxrows, columns)
        dict_table[lastd2+'dvolume'] = current[columns[diffV_col1]+str(diffV_row1)].value
        dict_table[lastd2+'denergy'] = current[columns[diffV_col1]+str(diffV_row1+1)].value
        dict_table[lastd2+'tolerance'] = current[columns[diffV_col1-1]+str(diffV_row1+2)].value
        dict_table[lastd2+'pass/fail'] = current[columns[diffV_col1-2]+str(diffV_row1+3)].value
        


    return dict_table

def searchstring(startrow, string1, string2, current, maxrows, columns):
    is_looping = True
    for row in range(startrow,maxrows):
        for col in range(0,20):
            if current[columns[col]+str(row)].value == string1 and current[columns[col+1]+str(row)].value == string2:
                #print(current[columns[col]+str(row)].value)
                #print(col, row)
                is_looping = False
                break

        if not is_looping:
            break # break out of outer loop
    return col+2,row
            

def count_max(wb):
    count = 0
    for row in wb: 
        if not all ([cell.value == None for cell in row]): 
            count += 1
    return count

#check as found or as left
def check_as(result):
    if 'f' in result or 'F' in result:
        if 'L' in result or 'l' in result:
            lastd = "al"
        else:
            lastd = "af"
    else:
        lastd = "al"
    return lastd

