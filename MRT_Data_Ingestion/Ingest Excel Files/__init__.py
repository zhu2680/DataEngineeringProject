import logging

import azure.functions as func

import azure.storage.blob


import io
from openpyxl import load_workbook
from string import ascii_uppercase as alp 
import numbers




def main(myblob: func.InputStream):
    logging.basicConfig(level = logging.INFO)
    logging.info(f"Python blob trigger function processed blob \n"
                 f"Name: {myblob.name}\n"
                 f"Blob Size: {myblob.length} bytes")
    columns = []    #build array to call columns from
    for i in alp: 
        columns.append(i)
    for i in alp: 
        for j in alp: 
            columns.append(i+j)

    #countABB = 1   # row counter for ABB
    #count = 1 # row counter for Barton
    logging.info('Loading Workbook...')
    if myblob.name.endswith('.xlsx'): 
        
        wb = load_workbook(filename = io.BytesIO(myblob.read()), data_only= True)
        current = wb.worksheets[0] #current file we are looking at 

        title = current['B3'].value

        logging.info("The title of the sheet is ABB/barton")
        if title != None and "ABB" in title: #check for ABB table
            try:
                logging.info("ABB")
                # this is our master file where we will write all info too
            
                ### Upload to Azure Blob Storage
                conString = "DefaultEndpointsProtocol=https;AccountName=mrtbartonabb3a1607b;AccountKey=awSIfw1YfTIclurJu6hrYrjjbgNMK1cTliTnEUv6VfFVKCfrmFAvzYx/m6ccV2ZnfBw2SaKORipp+AStQh5dwQ==;EndpointSuffix=core.windows.net"
                ##conString = "DefaultEndpointsProtocol=https;AccountName=mrtbartonabb3a1607b;AccountKey=eoJHu+6vJYYXEiv59LuHSouVaRxVL/XRcXkpVBQu4oVnKUY25Q1ZKgPb7LgbC/ax5jYTY5/r0DEY+AStkH/00w==;EndpointSuffix=core.windows.net"
                blob_service_client = azure.storage.blob.BlobServiceClient.from_connection_string(conString)
                    
                blob_client = blob_service_client.get_blob_client(container='abbtest', blob='MasterABB.xlsx')
                
                logging.info("downloading master ABB file...")
                with io.BytesIO() as input_blob:
                    blob_client.download_blob().readinto(input_blob)
                    input_blob.seek(0)
                    result = load_workbook(filename = input_blob, data_only= True)

                full_table = result.worksheets[0] #current file we are looking at 
                logging.info("finished loading ABB master file")


                #create the header and table dictionary
                dict = build_dictionary(current)
                dict2,count_d,count_t,count_s = build_dictionary_fail(current) 
                max_fails = max_fail(count_d,count_t,count_s)


                countABB = count_max(full_table)
                logging.info(f"The max number of non-empty row is: {countABB}")
                #if there is failed point for the report
                if max_fails >= 1: 
                    for i in range(0,max_fails): #Loop through the max number of failed points to record all failed points for a report
                        countABB += 1 
                        name1 = myblob.name[:-4]
                        full_table['A'+ str(countABB)].value= name1[4:]

                        for num, item in enumerate(dict, start=1): #always the same header data for failed points in the same report
                            full_table[columns[num]+ str(countABB)].value = dict[item]
                        
                        full_table['M'+ str(countABB)].value = dict2["pass/fail"] #This entry will be false
                        
                        #the code below fills the data for failed points
                        # e.g.
                        #if the report has 1 failed SAF and DAF, they will be appended to the same row
                        #if the report has 2 failed SAF and 1 DAF, an additional row with the same header data will be created to record the remaining failed SAF
                        if "SAFEntry"+str(i+1) in dict2: #failed SAF
                            full_table['N'+str(countABB)].value = dict2["SAFEntry"+str(i+1)]
                            full_table['O'+str(countABB)].value= dict2["SAFReading"+str(i+1)]
                            full_table['P'+str(countABB)].value = dict2["SAFError"+str(i+1)]
                        if "DAFEntry"+str(i+1) in dict2:#failed DAF
                            full_table['Q'+str(countABB)].value = dict2["DAFEntry"+str(i+1)]
                            full_table['R'+str(countABB)].value = dict2["DAFReading"+str(i+1)]
                            full_table['S'+str(countABB)].value = dict2["DAFError"+str(i+1)]

                        if "TAFEntry"+str(i+1) in dict2:#failed TAF
                            full_table['T'+str(countABB)].value = dict2["TAFEntry"+str(i+1)]
                            full_table['U'+str(countABB)].value = dict2["TAFReading"+str(i+1)]
                            full_table['V'+str(countABB)].value = dict2["TAFError"+str(i+1)]
                
                else:#if there is no failed points
                    countABB += 1 
                    name1 = myblob.name[:-4]
                    full_table['A'+ str(countABB)].value= name1[4:]
                    #regular entries for header, table data will be blank
                    for num, item in enumerate(dict, start=1):
                        full_table[columns[num]+ str(countABB)].value = dict[item]
                            
                full_table['M'+ str(countABB)].value = dict2["pass/fail"]  #this entry will be pass
                logging.info("reading info from ABB complete")

                    
                
                output = io.BytesIO()
                result.save(output)
                blob_client.upload_blob(output.getvalue(),overwrite=True)
                
                #test.save('C:\Users\mengyan_zhu\OneDrive - TransCanada Corporation\MRTResultFile\ABB.xlsx')
                logging.info("finish writing to a new ABB file")
            except Exception as e: 
                logging.error(f"An error has occured: {e}")
    
        
            
        else: 
            try:
                #VERSION 1
                if "Barton" in str(current['B3'].value):
                    logging.info("Barton1")
                    dictb = create_dictionary_v1head(current)
                    listb = create_dictionary_v1data(current)
                # VERSION 2
                elif "Barton" in str(current['A2'].value):
                    logging.info("Barton2")
                    dictb= create_dictionary_v2head(current)
                    listb = create_dictionary_v2data(current)
                #VERSION 3
                elif "Barton" in str(current['A1'].value):
                    logging.info("Barton3")
                    dictb= create_dictionary_v3head(current)
                    listb = create_dictionary_v3data(current)
                else: 
                    logging.info("ABB")

                    #continue
                logging.info("reading info from Barton complete")
                #open our file we will write into
               

                
                conString = "DefaultEndpointsProtocol=https;AccountName=mrtbartonabb3a1607b;AccountKey=awSIfw1YfTIclurJu6hrYrjjbgNMK1cTliTnEUv6VfFVKCfrmFAvzYx/m6ccV2ZnfBw2SaKORipp+AStQh5dwQ==;EndpointSuffix=core.windows.net"
                ##"DefaultEndpointsProtocol=https;AccountName=mrtbartonabb3a1607b;AccountKey=eoJHu+6vJYYXEiv59LuHSouVaRxVL/XRcXkpVBQu4oVnKUY25Q1ZKgPb7LgbC/ax5jYTY5/r0DEY+AStkH/00w==;EndpointSuffix=core.windows.net"
                blob_service_client = azure.storage.blob.BlobServiceClient.from_connection_string(conString)
                #blob_client = blob_service_client.get_blob_client(container='bartontest', blob='test2.xlsx')
                blob_client = blob_service_client.get_blob_client(container='bartontest', blob='MasterBarton.xlsx')
               
                
                #downloaded = blob_client.download_blob()
                #stream  = io.BytesIO(downloaded.readall())
                #result = load_workbook(filename = stream, data_only= True)
                
               
                with io.BytesIO() as input_blob:
                    blob_client.download_blob().readinto(input_blob)
                    input_blob.seek(0)
                    result = load_workbook(filename = input_blob, data_only= True)

                tableb1 = result.worksheets[0] #current file we are looking at 
                
                count= count_max(tableb1)
                
                logging.info(f"The max number of non-empty row is: {count}")
                
                
                
               
                #for each fail point in listb 
                for count1, dict in enumerate(listb, start=1):
                    count = count + 1    #initialized outside of for each file loop
                    
                    name1 = myblob.name[:-4]
                    tableb1['A'+ str(count)].value = name1[4:]

                    # for every dictionary(each fail point) in the list 
                    for num1, item in enumerate(dictb, start =1): 
                        # this will write our header data
                        #print("Current Column " + str(num1))
                        #print("current row "+ str(count))
                        tableb1[columns[num1]+ str(count)].value = dictb[item]
                    
                    for num3, data in enumerate(dict, start = 1):
                        # this will write our fail point data 
                        totalnum = num3 + num1
                       
                        tableb1[columns[totalnum]+ str(count)].value = dict[data]
               
                logging.info("Start writing to a new Barton file")
                
                
                output = io.BytesIO()
                
                result.save(output)
                blob_client.upload_blob(output.getvalue(),overwrite=True) 
                listb.clear()
            except Exception as e: 
                logging.error(f"An error in Barton has occured: {e}")
                        

    
    
#Barton Headers
def build_dictionary(current):
   
    dict_table = {}
   
    dict_table["date"] =current['D8'].value
    
    dict_table['stn.name']= current['D11'].value
    dict_table['stn.#'] = current['D12'].value

    dict_table["tech.name"] = current['D10'].value
    dict_table['run#'] = current['D13'].value
    dict_table['xfcmakemodel'] = current['D14'].value
    dict_table['xfcserial#']= current['D15'].value
    dict_table['rtdmakemodel']=current['D16'].value
    dict_table['rtdserial#'] = current['D17'].value
    dict_table['staticspan'] = current['F19'].value
    dict_table['differentialspan']=current['F20'].value
    
    
    
    return dict_table



#Check if the entries pass for each of the tables; Replacement of the rules for conditional formatting in excel
#purpose: identify the failed entry marked as red
def check_pass(entry, error, span, unit, minUS,maxUS, minMetric, maxMetric):
    if isinstance(entry, numbers.Number) and isinstance(error, numbers.Number):
    #entry != "N/A" and error != "N/A":
       
        if unit == "US Customary":
           
            if entry > span*0.1 and abs(error) > minUS:
               
                return True
            elif entry <= span*0.1 and abs(error) > maxUS:
                return True
            else:
                return False
        elif unit == "Metric":
            
            if entry > span*0.1 and abs(error) > minMetric:
                return True
            elif entry <= span*0.1 and abs(error) > maxMetric:
                return True
            else:
                return False
    return False       

#Record all the rows of the failed entries in a list for each of the table
#taking in one of the tables at a time and call the check_pass function to find out the failed entries
def failed_entries(current, mode):
    failed_entry = []
    unit = current['P19'].value     
    
    #conditional formatting for static
    if mode == "static":
        span = current['D23'].value
        
        for i in range(0,5):
            error = current['J'+ str(25+i)].value
            entry = current['D'+ str(25+i)].value
            fail = check_pass(entry, error, span, unit, 1.45038,0.290075001,10,2) 
            
            if fail:
                failed_entry.append(str(25+i))
    
    #conditional formatting for differential
    if mode == "differential":
        span = current['D32'].value
        for i in range(0,5):
            error = current['J'+ str(34+i)].value
            entry = current['D'+ str(34+i)].value
            fail = check_pass(entry, error, span, unit, 1.61,0.8, 0.4,0.2 ) 
            
            if fail:
                failed_entry.append(str(34+i))

     #conditional formatting for temperature
    if mode == "temperature":
        span = current['F41'].value
        for i in range(0,3):
            error = current['J'+ str(43+i)].value
            if span == "ºC" and abs(error)>1:
                failed_entry.append(str(43+i))
            
   
    return failed_entry #a list of failed row numbers
                
# build the dictionary of the failed points for each of the tables. Also record how many fails for each of the tables
def build_dictionary_fail(current):    
    dict_table = {}
    count_s = 0
    count_d = 0
    count_t = 0
    
    #static
    failed_point_s = failed_entries(current, "static")
    if failed_point_s: # If there is elements, add its data to dictionary 
        for i in failed_point_s:
            #count_s increments to distinguish each failed points and record the number of failed points for each table
            count_s += 1
            dict_table["SAFEntry"+str(count_s)] = current['D'+i].value  #e.g. SAFEntry1,SAFEntry2, ETC.
            dict_table["SAFReading"+str(count_s)] = current['G'+i].value
            dict_table["SAFError"+str(count_s)]=  current['J'+i].value
            

     
    #Differential
    failed_point_d = failed_entries(current, "differential")
    if failed_point_d:# If there is elements, add its data to dictionary 
        for i in failed_point_d:
            #count_d increments to distinguish each failed points and record the number of failed points for each table
            count_d +=1
            dict_table["DAFEntry"+str(count_d)] = current['D'+i].value #e.g. DAFEntry1,DAFEntry2, ETC.
            dict_table["DAFReading"+str(count_d)] = current['G'+i].value
            dict_table["DAFError"+str(count_d)]=  current['J'+i].value
            
    #Temperature
    failed_point_t = failed_entries(current, "temperature")
    if failed_point_t:# If there is elements, add its data to dictionary
        for i in failed_point_t:
            #count_t increments to distinguish each failed points and record the number of failed points for each table
            count_t += 1
            dict_table["TAFEntry"+str(count_t)] = current['D'+i].value #e.g. TAFEntry1,TAFEntry2, ETC.
            dict_table["TAFReading"+str(count_t)] = current['G'+i].value
            dict_table["TAFError"+str(count_t)]=  current['J'+i].value
            
    #if any of these lists is not blank, pass/fail will be fail
    if not failed_point_d and not failed_point_s and not failed_point_t:
        dict_table["pass/fail"] = "Pass"   
    else:
        dict_table["pass/fail"] = "Fail"   

    
    return dict_table,count_d,count_t,count_s

#check the maximum number of fails for each of the tables for each ABB. 
# max_fail will be the # of entries in the result excel for that ABB report
def max_fail(a,b,c):
    if a > b:
        if a > c:
            max_fail = a
        else:
            max_fail= c
    else:
        if b>c:
            max_fail= b
        else:
            max_fail = c

    return max_fail

#this fct gets values from merged cells 
def getMergedCellVal(sheet,  ref):
    cell = sheet[ref]
    if cell.coordinate in sheet.merged_cells: 
        for range_str in sheet.merged_cells: 
            if cell.coordinate in range_str: 
                start_cell= range_str.split(":")[0]
                merged_value = sheet[start_cell].value 
                return merged_value 
            return cell.value
    
#Header Data for Version 1 of Barton
def create_dictionary_v1head(current):
    dict_v1 = {}
    dict_v1['Date'] = getMergedCellVal(current, 'D7')
    dict_v1['Stn Name'] = getMergedCellVal(current, 'D9')
    dict_v1['Stn Number'] = getMergedCellVal(current, 'D10')
    dict_v1['FC Make/Model'] = getMergedCellVal(current, 'D11' )
    dict_v1['FC board Serial Num'] = getMergedCellVal(current, 'D12')
    dict_v1['Run Num'] = getMergedCellVal(current, 'D13')
    dict_v1['Static Make'] = getMergedCellVal(current, 'D14')
    dict_v1['Static Serial'] = getMergedCellVal(current, 'D15')
    dict_v1['Diff Make'] = getMergedCellVal(current, 'D16')
    dict_v1['Diff Serial'] = getMergedCellVal(current, 'D17')
    dict_v1['Temp Make'] = getMergedCellVal(current, 'D18')
    dict_v1['Temp Serial'] = getMergedCellVal(current, 'D19')
    dict_v1['Completed By'] = getMergedCellVal(current, 'L17')
    dict_v1['Static Span'] = current['E22'].value
    dict_v1['Voltage Allowance'] = current['E23'].value
    dict_v1['DP Span'] = current['E31'].value 
    dict_v1['Temp Span'] = current['E47'].value 
    dict_v1['Temp Allowance'] = current['E49'].value 
    dict_v1['Temp Low Range'] =current['H47'].value 
    dict_v1['Temp Upper Range'] = current['H48'].value
    return dict_v1
#Header Data for Version 2 of Barton
def create_dictionary_v2head(current):
    dict_v2 = {}
    dict_v2['Date'] = getMergedCellVal(current, 'C6')
    dict_v2['Stn Name'] = getMergedCellVal(current, 'C8')
    dict_v2['Stn Number'] = getMergedCellVal(current, 'C9')
    dict_v2['FC Make/Model'] = ""
    dict_v2['FC board Serial Num'] = ""
    dict_v2['Run Num'] = getMergedCellVal(current, 'C10')
    dict_v2['Static Make'] = ""
    dict_v2['Static Serial'] = getMergedCellVal(current, 'C11')
    dict_v2['Diff Make'] = ""
    dict_v2['Diff Serial'] = getMergedCellVal(current, 'C12')
    dict_v2['Temp Make'] = ""
    dict_v2['Temp Serial'] = getMergedCellVal(current, 'C13')
    dict_v2['Completed By'] = getMergedCellVal(current, 'K12')
    dict_v2['Static Span'] = current['D16'].value
    dict_v2['Voltage Allowance'] = current['D17'].value
    dict_v2['DP Span'] = current['D25'].value 
    dict_v2['Temp Span'] = current['D41'].value 
    dict_v2['Temp Allowance'] = current['D43'].value 
    dict_v2['Temp Low Range'] =current['G41'].value 
    dict_v2['Temp Upper Range'] = current['G42'].value
    return dict_v2
#Header Data for Version 3 of Barton
def create_dictionary_v3head(current):
    dict_v3 = {}
    dict_v3['Date'] = getMergedCellVal(current, 'C7')
    dict_v3['Stn Name'] = getMergedCellVal(current, 'C9')
    dict_v3['Stn Number'] = getMergedCellVal(current, 'C10')
    dict_v3['FC Make/Model'] = ""
    dict_v3['FC board Serial Num'] = ""
    dict_v3['Run Num'] = getMergedCellVal(current, 'C11')
    dict_v3['Static Make'] = ""
    dict_v3['Static Serial'] = getMergedCellVal(current, 'C12')
    dict_v3['Diff Make'] = ""
    dict_v3['Diff Serial'] = getMergedCellVal(current, 'C13')
    dict_v3['Temp Make'] = ""
    dict_v3['Temp Serial'] = ""
    dict_v3['Completed By'] = ""
    dict_v3['Static Span'] = current['C16'].value
    dict_v3['Voltage Allowance'] = current['C19'].value
    dict_v3['DP Span'] = current['C27'].value 
    dict_v3['Temp Span'] = "" 
    dict_v3['Temp Allowance'] = "" 
    dict_v3['Temp Low Range'] ="" 
    dict_v3['Temp Upper Range'] = ""
    return dict_v3

#Calibration Data for Version 1 of Barton
def create_dictionary_v1data(current): 
    # create list to append all the dictionarys too 
    dict_list = []
    dict_diff = {}
    dict_static = {}
    dict_temp = {}
    points = ['low', 'mid', 'high']

    #CHECK STATIC 
    for num, point in enumerate(points, start = 0): 
        row = 26+num
        # uncomment if you only want fail points 
        # if current['C'+ str(26+num)].value != None:      # if there is no test pressure dont create a fail point 
        #     if current['D'+ str(26+num)].value != None: # if there is no scanner pressure dont create a fail point 
        #      if current['H'+ str(26+num)].value == "Fail": 
        dict_static["Fail Type"] = "Static"
        dict_static["Fail Point"] = point
        dict_static["Test Pressure"] = current['C'+ str(row)].value
        dict_static["Scanner Pressure"] = current['D'+ str(row)].value 
        dict_static["Voltage (calc)"] = current['E'+ str(row)].value
        dict_static["Voltage (raw)"] = current['F'+ str(row)].value
        dict_static["Error"] = current['G'+ str(row)].value
        dict_static["Temp Target"] = ""
        dict_static["Temp Reading"] = ""
        dict_static["Temp Difference"] = ""
        dict_static["Result"] = current['H'+ str(row)].value
        dict_list.append(dict_static.copy())

    #CHECK DIFF 
    for num in range(0, 10): 
        row = 35+ num
        # uncomment if you only want fail points 
        # if current['C' +  str(row)].value != None:  # if there is no test pressure dont create a fail point 
        #     if current['D' +  str(row)].value != None:  # if there is no scanner pressure dont create a fail point
        #         if current['H' +  str(row)].value == "Fail":    # if says anything other than fail continue 
        dict_diff["Fail Type"] = "Differential"
        dict_diff["Fail Point"] = ""
        dict_diff["Test Pressure"] = current['C'+ str(row)].value
        dict_diff["Scanner Pressure"] = current['D'+ str(row)].value 
        dict_diff["Voltage (calc)"] = current['E'+ str(row)].value
        dict_diff["Voltage (raw)"] = current['F'+ str(row)].value
        dict_diff["Error"] = current['G'+ str(row)].value
        dict_diff["Temp Target"] = ""
        dict_diff["Temp Reading"] = ""
        dict_diff["Temp Difference"] = ""
        dict_diff["Result"] = current['H'+ str(row)].value
        dict_list.append(dict_diff.copy())
                    
    #CHECK TEMP 
    for num, point in enumerate(points, start = 0):
        row = 52+num
        # uncomment if you only want fail points 
        # if current['C' +  str(52 + num)] != None:   # if there is no test temp dont create a fail point
        #     if current['D' +  str(52 + num)] != None:   # if there is no reading dont create a fail point
        #         if  getMergedCellVal(current, 'G' +  str(52 + num)) == "Fail":
        dict_temp["Fail Type"] = "Temp"
        dict_temp["Fail Point"] = point
        dict_temp["Test Pressure"] = ""
        dict_temp["Scanner Pressure"] = "" 
        dict_temp["Voltage (calc)"] = ""
        dict_temp["Voltage (raw)"] = ""
        dict_temp["Error"] = ""
        dict_temp["Temp Target"] = current['C' +  str(row)].value
        dict_temp["Temp Reading"] = current['D' +  str(row)].value
        dict_temp["Temp Difference"] = getMergedCellVal(current, 'E' +  str(row))
        dict_temp["Result"] = getMergedCellVal(current, 'G' + str(row))
        dict_list.append(dict_temp.copy())

    return dict_list
#Calibration Data for Version 2 of Barton
def create_dictionary_v2data(current): 
    # create list to append all the dictionarys too 
    dict_list = []
    dict_diff = {}
    dict_static = {}
    dict_temp = {}
    points = ['low', 'mid', 'high']

    #CHECK STATIC 
    for num, point in enumerate(points, start = 0):
        row = 20 + num
        #uncomment if you only want fail points 
        # if current['B'+ str(20+num)].value != None:      # if there is no test pressure dont create a fail point 
        #     if current['C'+ str(20+num)].value != None: # if there is no scanner pressure dont create a fail point 
        #      if current['G'+ str(20+num)].value == "Fail": 
        dict_static["Fail Type"] = "Static"
        dict_static["Fail Point"] = point
        dict_static["Test Pressure"] = current['B'+ str(row)].value
        dict_static["Scanner Pressure"] = current['C'+ str(row)].value 
        dict_static["Voltage (calc)"] = current['D'+ str(row)].value
        dict_static["Voltage (raw)"] = current['E'+ str(row)].value
        dict_static["Error"] = current['F'+ str(row)].value
        dict_static["Temp Target"] = ""
        dict_static["Temp Reading"] = ""
        dict_static["Temp Difference"] = ""
        dict_static["Result"] = current['G'+ str(row)].value
        dict_list.append(dict_static.copy())

    #CHECK DIFF 
    for num in range(0, 10): 
        row = 29+ num
        #uncomment if you only want fail points 
        # if current['B' +  str(row)].value != None:  # if there is no test pressure dont create a fail point 
        #     if current['C' +  str(row)].value != None:  # if there is no scanner pressure dont create a fail point
        #         if current['G' +  str(row)].value == "Fail":    # if says anything other than fail continue 
        dict_diff["Fail Type"] = "Differential"
        dict_diff["Fail Point"] = ""
        dict_diff["Test Pressure"] = current['B'+ str(row)].value
        dict_diff["Scanner Pressure"] = current['C'+ str(row)].value 
        dict_diff["Voltage (calc)"] = current['D'+ str(row)].value
        dict_diff["Voltage (raw)"] = current['E'+ str(row)].value
        dict_diff["Error"] = current['F'+ str(row)].value
        dict_diff["Temp Target"] = ""
        dict_diff["Temp Reading"] = ""
        dict_diff["Temp Difference"] = ""
        dict_diff["Result"] = current['G'+ str(row)].value
        dict_list.append(dict_diff.copy())
                    
    #CHECK TEMP 
    for num, point in enumerate(points, start = 0):
        row = 46+num
        #uncomment if you only want fail points 
        # if current['B' +  str(46 + num)] != None:   # if there is no test temp dont create a fail point
        #     if current['C' +  str(46 + num)] != None:   # if there is no reading dont create a fail point
        #         if  getMergedCellVal(current, 'F' +  str(46 + num)) == "Fail":
        dict_temp["Fail Type"] = "Temp"
        dict_temp["Fail Point"] = point
        dict_temp["Test Pressure"] = ""
        dict_temp["Scanner Pressure"] = "" 
        dict_temp["Voltage (calc)"] = ""
        dict_temp["Voltage (raw)"] = ""
        dict_temp["Error"] = ""
        dict_temp["Temp Target"] = current['B' +  str(row)].value
        dict_temp["Temp Reading"] = current['C' +  str(row)].value
        dict_temp["Temp Difference"] = getMergedCellVal(current, 'D' +  str(row))
        dict_temp["Result"] = getMergedCellVal(current, 'F' + str(row))
        dict_list.append(dict_temp.copy())

    return dict_list
#Calibration Data For Version 3 of Barton 
def create_dictionary_v3data(current): 
     #note: there is no temperature category for v3
    # create list to append all the dictionarys too 
    dict_list = []
    dict_diff = {}
    dict_static = {}
    points = ['low', 'mid', 'high']

    #CHECK STATIC 
    for num, point in enumerate(points, start = 0):
        row = 22+num 
        # uncomment if you only want fail points
        # if current['C'+ str(22+num)].value != None:      # if there is no test pressure dont create a fail point 
        #     if current['E'+ str(22+num)].value != None: # if there is no scanner pressure dont create a fail point 
        #      if current['H'+ str(22+num)].value == "Fail": 
        dict_static["Fail Type"] = "Static"
        dict_static["Fail Point"] = point
        dict_static["Test Pressure"] = current['C'+ str(row)].value
        dict_static["Scanner Pressure"] = ""
        dict_static["Voltage (calc)"] = current['D'+ str(row)].value
        dict_static["Voltage (raw)"] = current['E'+ str(row)].value
        dict_static["Error"] = current['G'+ str(row)].value
        dict_static["Temp Target"] = ""
        dict_static["Temp Reading"] = ""
        dict_static["Temp Difference"] = ""
        dict_static["Result"] = current['H'+ str(row)].value
        dict_list.append(dict_static.copy())

    #CHECK DIFF 
    for num in range(0, 10): 
        row = 30+ num
        # uncomment if you only want fail points
        # if current['C' +  str(row)].value != None:  # if there is no test pressure dont create a fail point 
        #     if current['E' +  str(row)].value != None:  # if there is no scanner pressure dont create a fail point
        #         if current['H' +  str(row)].value == "Fail":    # if says anything other than fail continue 
        dict_diff["Fail Type"] = "Differential"
        dict_diff["Fail Point"] = ""
        dict_diff["Test Pressure"] = current['C'+ str(row)].value
        dict_diff["Scanner Pressure"] = ""
        dict_diff["Voltage (calc)"] = current['D'+ str(row)].value
        dict_diff["Voltage (raw)"] = current['E'+ str(row)].value
        dict_diff["Error"] = current['G'+ str(row)].value
        dict_diff["Temp Target"] = ""
        dict_diff["Temp Reading"] = ""
        dict_diff["Temp Difference"] = ""
        dict_diff["Result"] = current['H'+ str(row)].value
        dict_list.append(dict_diff.copy())
                    
    return dict_list


def count_max(wb):
    count = 0
    for row in wb: 
        if not all ([cell.value == None for cell in row]): 
            count += 1
    return count